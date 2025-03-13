#!/usr/bin/env python
# -*- coding: utf-8 -*-

"""
Модуль виджета отчетов для WDM_V12.
Обеспечивает интерфейс для создания, предварительного просмотра
и экспорта отчетов мониторинга в различных форматах.
"""

import os
import csv
import json
from datetime import datetime, timedelta
from pathlib import Path

from PyQt6.QtWidgets import (
    QWidget, QVBoxLayout, QLabel, QPushButton, QComboBox, QHBoxLayout,
    QGroupBox, QGridLayout, QDateEdit, QCheckBox, QFileDialog, QMessageBox,
    QTableWidget, QTableWidgetItem, QHeaderView, QTextEdit, QSplitter,
    QTabWidget, QScrollArea, QFrame, QStackedWidget, QRadioButton, QSpinBox,
    QPrintDialog
)
from PyQt6.QtCore import Qt, QDate, QDateTime, QSize
from PyQt6.QtGui import QIcon, QFont, QTextDocument
from PyQt6.QtPrintSupport import QPrinter

from utils.logger import get_module_logger, log_exception
from core.settings import Settings
from reports.report_generator import ReportGenerator
from utils.decorators import handle_errors


class ReportsWidget(QWidget):
    """
    Виджет для работы с отчетами.
    Предоставляет интерфейс для создания и экспорта отчетов мониторинга.
    """
    
    def __init__(self, app_context, parent=None):
        """
        Инициализация виджета отчетов
        
        Args:
            app_context: Контекст приложения
            parent: Родительский виджет
        """
        super().__init__(parent)
        
        self.logger = get_module_logger('ui.reports')
        self.logger.debug("Инициализация виджета отчетов")
        
        self.app_context = app_context
        self.settings = Settings()
        
        # Текущий отчет
        self.current_report = {
            'type': 'sites',
            'format': 'html',
            'date_from': QDate.currentDate().addDays(-7),
            'date_to': QDate.currentDate(),
            'content': None,
            'data': None
        }
        
        # Создание UI
        self._init_ui()
        
        self.logger.debug("Виджет отчетов инициализирован")
    
    def _init_ui(self):
        """Инициализация пользовательского интерфейса"""
        # Основной макет
        main_layout = QVBoxLayout(self)
        
        # Разделитель для размещения параметров отчета и предпросмотра
        splitter = QSplitter(Qt.Orientation.Horizontal)
        
        # Левая панель (параметры отчета)
        params_widget = QWidget()
        params_layout = QVBoxLayout(params_widget)
        
        # Группа параметров отчета
        report_params_group = QGroupBox("Параметры отчета")
        report_params_layout = QGridLayout(report_params_group)
        
        # Тип отчета
        report_params_layout.addWidget(QLabel("Тип отчета:"), 0, 0)
        self.report_type_combo = QComboBox()
        self.report_type_combo.addItems([
            "По сайтам", 
            "По изменениям", 
            "По ошибкам",
            "Статистика"
        ])
        self.report_type_combo.currentIndexChanged.connect(self._on_report_type_changed)
        report_params_layout.addWidget(self.report_type_combo, 0, 1, 1, 2)
        
        # Период
        report_params_layout.addWidget(QLabel("Период:"), 1, 0)
        
        # От даты
        self.date_from = QDateEdit()
        self.date_from.setCalendarPopup(True)
        self.date_from.setDate(self.current_report['date_from'])
        report_params_layout.addWidget(self.date_from, 1, 1)
        
        # До даты
        self.date_to = QDateEdit()
        self.date_to.setCalendarPopup(True)
        self.date_to.setDate(self.current_report['date_to'])
        report_params_layout.addWidget(self.date_to, 1, 2)
        
        # Добавляем группу параметров в левую панель
        params_layout.addWidget(report_params_group)
        
        # Группа форматирования отчета
        formatting_group = QGroupBox("Форматирование")
        formatting_layout = QGridLayout(formatting_group)
        
        # Формат отчета
        formatting_layout.addWidget(QLabel("Формат:"), 0, 0)
        self.format_combo = QComboBox()
        self.format_combo.addItems(["HTML", "CSV", "XLSX", "PDF"])
        self.format_combo.currentIndexChanged.connect(self._on_format_changed)
        formatting_layout.addWidget(self.format_combo, 0, 1, 1, 2)
        
        # Включить диаграммы
        self.include_charts = QCheckBox("Включить диаграммы")
        self.include_charts.setChecked(True)
        formatting_layout.addWidget(self.include_charts, 1, 0, 1, 3)
        
        # Включить статистику
        self.include_stats = QCheckBox("Включить статистику")
        self.include_stats.setChecked(True)
        formatting_layout.addWidget(self.include_stats, 2, 0, 1, 3)
        
        # Добавляем группу форматирования в левую панель
        params_layout.addWidget(formatting_group)
        
        # Кнопки действий
        actions_layout = QHBoxLayout()
        
        # Кнопка генерации отчета
        self.generate_button = QPushButton("Создать отчет")
        self.generate_button.setIcon(QIcon("resources/icons/report.png"))
        self.generate_button.clicked.connect(self._on_generate_report)
        actions_layout.addWidget(self.generate_button)
        
        # Кнопка экспорта
        self.export_button = QPushButton("Экспорт")
        self.export_button.setIcon(QIcon("resources/icons/export.png"))
        self.export_button.clicked.connect(self._on_export_report)
        self.export_button.setEnabled(False)  # Изначально отключена
        actions_layout.addWidget(self.export_button)
        
        # Кнопка печати
        self.print_button = QPushButton("Печать")
        self.print_button.setIcon(QIcon("resources/icons/print.png"))
        self.print_button.clicked.connect(self._on_print_report)
        self.print_button.setEnabled(False)  # Изначально отключена
        actions_layout.addWidget(self.print_button)
        
        # Добавляем кнопки в левую панель
        params_layout.addLayout(actions_layout)
        
        # Добавляем растяжитель для выравнивания
        params_layout.addStretch()
        
        # Добавляем левую панель в разделитель
        splitter.addWidget(params_widget)
        
        # Правая панель (предпросмотр отчета)
        preview_widget = QWidget()
        preview_layout = QVBoxLayout(preview_widget)
        
        # Заголовок предпросмотра
        preview_layout.addWidget(QLabel("Предварительный просмотр отчета"))
        
        # Область предпросмотра
        self.preview_area = QStackedWidget()
        
        # Страница для HTML-предпросмотра
        self.html_preview = QTextEdit()
        self.html_preview.setReadOnly(True)
        self.preview_area.addWidget(self.html_preview)
        
        # Страница для табличного предпросмотра
        self.table_preview = QTableWidget()
        self.table_preview.setAlternatingRowColors(True)
        self.preview_area.addWidget(self.table_preview)
        
        # Добавляем область предпросмотра в правую панель
        preview_layout.addWidget(self.preview_area)
        
        # Добавляем правую панель в разделитель
        splitter.addWidget(preview_widget)
        
        # Устанавливаем соотношение размеров
        splitter.setSizes([300, 700])
        
        # Добавляем разделитель в основной макет
        main_layout.addWidget(splitter)
    
    @handle_errors(error_msg="Ошибка при обновлении отчетов")
    def refresh(self):
        """Обновление списка отчетов"""
        try:
            # Получаем список отчетов
            reports = self.app_context.get_reports()
            
            # Очищаем таблицу
            self.table_preview.setRowCount(0)
            
            # Заполняем таблицу
            for row, report in enumerate(reports):
                self.table_preview.insertRow(row)
                
                # Тип отчета
                report_type = report.get('type', '')
                type_map = {
                    'sites': "Сайты",
                    'changes': "Изменения",
                    'errors': "Ошибки",
                    'stats': "Статистика"
                }
                type_text = type_map.get(report_type, report_type)
                type_item = QTableWidgetItem(type_text)
                self.table_preview.setItem(row, 0, type_item)
                
                # Дата создания
                timestamp = report.get('created_at')
                date_item = QTableWidgetItem(format_timestamp(timestamp))
                self.table_preview.setItem(row, 1, date_item)
                
                # Период
                date_from = report.get('date_from')
                date_to = report.get('date_to')
                period_text = f"{format_timestamp(date_from)} - {format_timestamp(date_to)}"
                period_item = QTableWidgetItem(period_text)
                self.table_preview.setItem(row, 2, period_item)
                
                # Размер файла
                file_path = report.get('file_path', '')
                if os.path.exists(file_path):
                    size = os.path.getsize(file_path)
                    if size < 1024:
                        size_text = f"{size} B"
                    elif size < 1024 * 1024:
                        size_text = f"{size/1024:.1f} KB"
                    else:
                        size_text = f"{size/(1024*1024):.1f} MB"
                else:
                    size_text = "Файл не найден"
                
                size_item = QTableWidgetItem(size_text)
                self.table_preview.setItem(row, 3, size_item)
                
                # Статус
                status = report.get('status', '')
                status_map = {
                    'completed': "Завершен",
                    'failed': "Ошибка",
                    'pending': "В процессе"
                }
                status_text = status_map.get(status, status)
                status_item = QTableWidgetItem(status_text)
                
                # Устанавливаем цвет статуса
                status_item.setForeground(get_status_color(status))
                self.table_preview.setItem(row, 4, status_item)
        except Exception as e:
            self.logger.error(f"Ошибка при обновлении списка отчетов: {e}")
            log_exception(self.logger, "Ошибка обновления списка отчетов")
            if hasattr(self.parent, "show_message"):
                self.parent.show_message("Ошибка", f"Не удалось обновить список отчетов: {e}", QMessageBox.Icon.Critical)
    
    def _on_report_type_changed(self, index):
        """
        Обработчик изменения типа отчета
        
        Args:
            index: Индекс выбранного типа отчета
        """
        self.logger.debug(f"Изменен тип отчета на {self.report_type_combo.currentText()}")
        
        # Обновляем текущий отчет
        report_types = ['sites', 'changes', 'errors', 'stats']
        if index < len(report_types):
            self.current_report['type'] = report_types[index]
        
        # Сбрасываем текущий отчет
        self.current_report['content'] = None
        self.current_report['data'] = None
        
        # Отключаем кнопки экспорта и печати
        self.export_button.setEnabled(False)
        self.print_button.setEnabled(False)
    
    def _on_format_changed(self, index):
        """
        Обработчик изменения формата отчета
        
        Args:
            index: Индекс выбранного формата
        """
        self.logger.debug(f"Изменен формат отчета на {self.format_combo.currentText()}")
        
        # Обновляем текущий отчет
        report_formats = ['html', 'csv', 'xlsx', 'pdf']
        if index < len(report_formats):
            self.current_report['format'] = report_formats[index]
        
        # Включаем/отключаем чекбоксы в зависимости от формата
        include_charts = self.current_report['format'] in ['html', 'pdf', 'xlsx']
        self.include_charts.setEnabled(include_charts)
        
        # Обновляем предпросмотр, если есть данные
        if self.current_report['data']:
            self._update_preview()
    
    def _on_generate_report(self):
        """Обработчик генерации отчета"""
        self.logger.debug("Вызван метод генерации отчета")
        
        try:
            # Получаем параметры отчета
            report_type = self.current_report['type']
            date_from = self.date_from.date().toPyDate()
            date_to = self.date_to.date().toPyDate()
            
            # Проверяем корректность дат
            if date_from > date_to:
                QMessageBox.warning(self, "Предупреждение", "Дата начала периода должна быть меньше даты окончания")
                return
            
            # Обновляем параметры отчета
            self.current_report['date_from'] = self.date_from.date()
            self.current_report['date_to'] = self.date_to.date()
            
            # Создаем генератор отчетов
            report_generator = ReportGenerator(self.app_context)
            
            # Генерируем отчет в зависимости от типа
            if report_type == 'sites':
                report_data = report_generator.generate_sites_report(date_from, date_to)
            elif report_type == 'changes':
                report_data = report_generator.generate_changes_report(date_from, date_to)
            elif report_type == 'errors':
                report_data = report_generator.generate_errors_report(date_from, date_to)
            elif report_type == 'stats':
                report_data = report_generator.generate_stats_report(date_from, date_to)
            else:
                raise ValueError(f"Неизвестный тип отчета: {report_type}")
            
            # Сохраняем данные отчета
            self.current_report['data'] = report_data
            
            # Форматируем отчет в HTML
            self.current_report['content'] = report_generator.format_report_html(report_data)
            
            # Обновляем предпросмотр
            self._update_preview()
            
            # Включаем кнопки экспорта и печати
            self.export_button.setEnabled(True)
            self.print_button.setEnabled(True)
            
            self.logger.info(f"Отчет {report_type} успешно сгенерирован")
        
        except Exception as e:
            self.logger.error(f"Ошибка при генерации отчета: {e}")
            log_exception(self.logger, "Ошибка генерации отчета")
            QMessageBox.critical(self, "Ошибка", f"Не удалось сгенерировать отчет: {e}")
    
    def _on_export_report(self):
        """Обработчик экспорта отчета"""
        self.logger.debug(f"Вызван метод экспорта отчета в формате {self.format_combo.currentText()}")
        
        try:
            # Получаем формат отчета
            report_format = self.current_report['format']
            
            # Определяем фильтр для файлового диалога
            format_filters = {
                'html': "HTML (*.html)",
                'csv': "CSV (*.csv)",
                'xlsx': "Excel (*.xlsx)",
                'pdf': "PDF (*.pdf)"
            }
            
            # Получаем фильтр и расширение
            file_filter = format_filters.get(report_format, "Все файлы (*.*)")
            extension = f".{report_format}"
            
            # Открываем диалог сохранения файла
            file_path, _ = QFileDialog.getSaveFileName(
                self,
                "Сохранить отчет",
                f"report_{self.current_report['type']}_{datetime.date.today().strftime('%Y%m%d')}{extension}",
                file_filter
            )
            
            if not file_path:
                return
            
            # Проверяем и при необходимости добавляем расширение
            if not file_path.endswith(extension):
                file_path += extension
            
            # Экспортируем отчет в зависимости от формата
            if report_format == 'html':
                self._export_html(file_path)
            elif report_format == 'csv':
                self._export_csv(file_path)
            elif report_format == 'xlsx':
                try:
                    # Пробуем импортировать openpyxl
                    import openpyxl
                    self._export_xlsx(file_path)
                except ImportError:
                    self.logger.error("Библиотека openpyxl не установлена")
                    QMessageBox.warning(
                        self,
                        "Отсутствует библиотека",
                        "Для экспорта в Excel требуется библиотека openpyxl.\n"
                        "Установите ее командой: pip install openpyxl"
                    )
            elif report_format == 'pdf':
                self._export_pdf(file_path)
            else:
                raise ValueError(f"Неизвестный формат отчета: {report_format}")
            
            # Сообщение об успешном экспорте
            if report_format in ['html', 'csv', 'xlsx', 'pdf'] and os.path.exists(file_path):
                self.logger.info(f"Отчет успешно экспортирован в {file_path}")
                QMessageBox.information(self, "Успех", f"Отчет успешно экспортирован в {file_path}")
        
        except Exception as e:
            self.logger.error(f"Ошибка при экспорте отчета: {e}")
            log_exception(self.logger, "Ошибка экспорта отчета")
            QMessageBox.critical(self, "Ошибка", f"Не удалось экспортировать отчет: {e}")
    
    def _export_html(self, file_path):
        """
        Экспорт отчета в HTML-формат
        
        Args:
            file_path: Путь для сохранения файла
        """
        self.logger.debug(f"Экспорт отчета в HTML-формат: {file_path}")
        
        # Проверяем наличие HTML-контента
        if not self.current_report['content']:
            # Обновляем HTML-контент
            self._update_html_preview()
        
        # Сохраняем HTML-контент в файл
        try:
            with open(file_path, 'w', encoding='utf-8') as f:
                f.write(self.current_report['content'])
        except Exception as e:
            self.logger.error(f"Ошибка при сохранении HTML-файла: {e}")
            raise
    
    def _export_csv(self, file_path):
        """
        Экспорт отчета в CSV-формат
        
        Args:
            file_path: Путь для сохранения файла
        """
        self.logger.debug(f"Экспорт отчета в CSV-формат: {file_path}")
        
        try:
            # Получаем данные отчета
            report_data = self.current_report['data']
            report_type = report_data['type']
            
            with open(file_path, 'w', newline='', encoding='utf-8') as csvfile:
                writer = csv.writer(csvfile)
                
                if report_type == 'sites':
                    # Заголовки для отчета по сайтам
                    headers = ['Название', 'URL', 'Группа', 'Статус', 'Последняя проверка', 
                             'Последнее изменение', 'Изменений', 'Ошибок']
                    writer.writerow(headers)
                    
                    # Записываем данные о сайтах
                    for site in report_data['sites']:
                        row = [
                            site['name'],
                            site['url'],
                            site['group_name'] or 'Без группы',
                            site['status'],
                            site['last_check'].strftime('%d.%m.%Y %H:%M') if site['last_check'] else '-',
                            site['last_change'].strftime('%d.%m.%Y %H:%M') if site['last_change'] else '-',
                            site['changes_count'],
                            site['errors_count']
                        ]
                        writer.writerow(row)
                
                elif report_type == 'changes':
                    # Заголовки для отчета по изменениям
                    headers = ['Сайт', 'Дата', 'Процент изменений', 'Статус', 'Проверил', 'Комментарий']
                    writer.writerow(headers)
                    
                    # Записываем данные об изменениях
                    for change in report_data['changes']:
                        row = [
                            change['site_name'],
                            change['timestamp'].strftime('%d.%m.%Y %H:%M'),
                            f"{change['diff_percent']:.2f}%",
                            change['status'],
                            change['reviewed_by'] or '-',
                            change['notes'] or '-'
                        ]
                        writer.writerow(row)
                
                elif report_type == 'errors':
                    # Заголовки для отчета по ошибкам
                    headers = ['Сайт', 'Дата', 'Сообщение об ошибке', 'Группа']
                    writer.writerow(headers)
                    
                    # Записываем данные об ошибках
                    for error in report_data['errors']:
                        row = [
                            error['name'],
                            error['error_time'].strftime('%d.%m.%Y %H:%M'),
                            error['error_message'],
                            error['group_name'] or '-'
                        ]
                        writer.writerow(row)
                
                elif report_type == 'stats':
                    # Записываем общую статистику
                    writer.writerow(['Общая статистика'])
                    writer.writerow(['Всего сайтов', report_data['sites_stats']['total_sites']])
                    writer.writerow(['Активных сайтов', report_data['sites_stats']['active_sites']])
                    writer.writerow(['Средний интервал проверки (мин)', 
                                  f"{report_data['sites_stats']['avg_check_interval'] / 60:.1f}"])
                    
                    writer.writerow([])  # Пустая строка
                    
                    # Статистика изменений
                    writer.writerow(['Статистика изменений'])
                    writer.writerow(['Всего изменений', report_data['changes_stats']['total_changes']])
                    writer.writerow(['Сайтов с изменениями', report_data['changes_stats']['sites_with_changes']])
                    writer.writerow(['Средний процент изменений', 
                                  f"{report_data['changes_stats']['avg_diff_percent']:.2f}%"])
                    
                    writer.writerow([])  # Пустая строка
                    
                    # Статистика ошибок
                    writer.writerow(['Статистика ошибок'])
                    writer.writerow(['Всего ошибок', report_data['errors_stats']['total_errors']])
                    writer.writerow(['Сайтов с ошибками', report_data['errors_stats']['sites_with_errors']])
        
        except Exception as e:
            self.logger.error(f"Ошибка при экспорте в CSV: {e}")
            log_exception(self.logger, "Ошибка экспорта в CSV")
            raise
    
    def _export_xlsx(self, file_path):
        """
        Экспорт отчета в Excel-формат
        
        Args:
            file_path: Путь для сохранения файла
        """
        self.logger.debug(f"Экспорт отчета в Excel-формат: {file_path}")
        
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment
            from openpyxl.utils import get_column_letter
            
            # Создаем новую книгу
            wb = openpyxl.Workbook()
            ws = wb.active
            
            # Получаем данные отчета
            report_data = self.current_report['data']
            report_type = report_data['type']
            
            # Стили
            header_font = Font(bold=True)
            header_fill = PatternFill(start_color='F0F0F0', end_color='F0F0F0', fill_type='solid')
            
            if report_type == 'sites':
                ws.title = 'Сайты'
                
                # Заголовки
                headers = ['Название', 'URL', 'Группа', 'Статус', 'Последняя проверка', 
                         'Последнее изменение', 'Изменений', 'Ошибок']
                
                for col, header in enumerate(headers, 1):
                    cell = ws.cell(row=1, column=col, value=header)
                    cell.font = header_font
                    cell.fill = header_fill
                
                # Данные
                for row, site in enumerate(report_data['sites'], 2):
                    ws.cell(row=row, column=1, value=site['name'])
                    ws.cell(row=row, column=2, value=site['url'])
                    ws.cell(row=row, column=3, value=site['group_name'] or 'Без группы')
                    ws.cell(row=row, column=4, value=site['status'])
                    ws.cell(row=row, column=5, value=site['last_check'].strftime('%d.%m.%Y %H:%M') if site['last_check'] else '-')
                    ws.cell(row=row, column=6, value=site['last_change'].strftime('%d.%m.%Y %H:%M') if site['last_change'] else '-')
                    ws.cell(row=row, column=7, value=site['changes_count'])
                    ws.cell(row=row, column=8, value=site['errors_count'])
            
            elif report_type == 'changes':
                ws.title = 'Изменения'
                
                # Заголовки
                headers = ['Сайт', 'Дата', 'Процент изменений', 'Статус', 'Проверил', 'Комментарий']
            
                for col, header in enumerate(headers, 1):
                    cell = ws.cell(row=1, column=col, value=header)
                    cell.font = header_font
                    cell.fill = header_fill
            
            # Данные
                for row, change in enumerate(report_data['changes'], 2):
                    ws.cell(row=row, column=1, value=change['site_name'])
                    ws.cell(row=row, column=2, value=change['timestamp'].strftime('%d.%m.%Y %H:%M'))
                    ws.cell(row=row, column=3, value=f"{change['diff_percent']:.2f}%")
                    ws.cell(row=row, column=4, value=change['status'])
                    ws.cell(row=row, column=5, value=change['reviewed_by'] or '-')
                    ws.cell(row=row, column=6, value=change['notes'] or '-')
            
            elif report_type == 'errors':
                ws.title = 'Ошибки'
                
                # Заголовки
                headers = ['Сайт', 'Дата', 'Сообщение об ошибке', 'Группа']
            
                for col, header in enumerate(headers, 1):
                    cell = ws.cell(row=1, column=col, value=header)
                    cell.font = header_font
                    cell.fill = header_fill
            
            # Данные
                for row, error in enumerate(report_data['errors'], 2):
                    ws.cell(row=row, column=1, value=error['name'])
                    ws.cell(row=row, column=2, value=error['error_time'].strftime('%d.%m.%Y %H:%M'))
                    ws.cell(row=row, column=3, value=error['error_message'])
                    ws.cell(row=row, column=4, value=error['group_name'] or '-')
            
            elif report_type == 'stats':
                ws.title = 'Статистика'
                
                current_row = 1
                
                # Общая статистика
                ws.cell(row=current_row, column=1, value='Общая статистика').font = header_font
                current_row += 1
                
                stats = [
                    ('Всего сайтов', report_data['sites_stats']['total_sites']),
                    ('Активных сайтов', report_data['sites_stats']['active_sites']),
                    ('Средний интервал проверки (мин)', f"{report_data['sites_stats']['avg_check_interval'] / 60:.1f}")
                ]
                
                for stat in stats:
                    ws.cell(row=current_row, column=1, value=stat[0])
                    ws.cell(row=current_row, column=2, value=stat[1])
                    current_row += 1
                
                current_row += 1
                
                # Статистика изменений
                ws.cell(row=current_row, column=1, value='Статистика изменений').font = header_font
                current_row += 1
                
                changes_stats = [
                    ('Всего изменений', report_data['changes_stats']['total_changes']),
                    ('Сайтов с изменениями', report_data['changes_stats']['sites_with_changes']),
                    ('Средний процент изменений', f"{report_data['changes_stats']['avg_diff_percent']:.2f}%"),
                    ('Максимальный процент изменений', f"{report_data['changes_stats']['max_diff_percent']:.2f}%")
                ]
                
                for stat in changes_stats:
                    ws.cell(row=current_row, column=1, value=stat[0])
                    ws.cell(row=current_row, column=2, value=stat[1])
                    current_row += 1
                
                current_row += 1
                
                # Статистика ошибок
                ws.cell(row=current_row, column=1, value='Статистика ошибок').font = header_font
                current_row += 1
                
                errors_stats = [
                    ('Всего ошибок', report_data['errors_stats']['total_errors']),
                    ('Сайтов с ошибками', report_data['errors_stats']['sites_with_errors'])
                ]
                
                for stat in errors_stats:
                    ws.cell(row=current_row, column=1, value=stat[0])
                    ws.cell(row=current_row, column=2, value=stat[1])
                    current_row += 1
            
            # Автоматическая ширина столбцов
            for column in ws.columns:
                max_length = 0
                column = [cell for cell in column]
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = (max_length + 2)
                ws.column_dimensions[get_column_letter(column[0].column)].width = adjusted_width
            
            # Сохраняем файл
            wb.save(file_path)
        
        except Exception as e:
            self.logger.error(f"Ошибка при экспорте в Excel: {e}")
            log_exception(self.logger, "Ошибка экспорта в Excel")
            raise
    
    def _export_pdf(self, file_path):
        """
        Экспорт отчета в PDF-формат
        
        Args:
            file_path: Путь для сохранения файла
        """
        self.logger.debug(f"Экспорт отчета в PDF-формат: {file_path}")
        
        try:
            # Создаем принтер для PDF
            printer = QPrinter(QPrinter.PrinterMode.HighResolution)
            printer.setOutputFormat(QPrinter.OutputFormat.PdfFormat)
            printer.setOutputFileName(file_path)
            printer.setPageSize(QPrinter.PageSize.A4)
            
            # Создаем документ
            document = QTextDocument()
            document.setHtml(self.current_report['content'])
            
            # Печатаем в PDF
            document.print(printer)
        
        except Exception as e:
            self.logger.error(f"Ошибка при экспорте в PDF: {e}")
            log_exception(self.logger, "Ошибка экспорта в PDF")
            raise
    
    def _update_preview(self):
        """Обновление предпросмотра отчета"""
        self.logger.debug("Обновление предпросмотра отчета")
        
        # Проверяем наличие данных
        if not self.current_report['data']:
            return
        
        # Обновляем предпросмотр в зависимости от формата
        report_format = self.current_report['format']
        
        if report_format == 'html':
            self._update_html_preview()
            self.preview_area.setCurrentIndex(0)  # Показываем HTML превью
        else:
            self._update_table_preview()
            self.preview_area.setCurrentIndex(1)  # Показываем табличное превью
    
    def _update_html_preview(self):
        """Обновление HTML-предпросмотра"""
        self.logger.debug("Обновление HTML-предпросмотра")
        
        # Получаем тип отчета и данные
        report_type = self.current_report['type']
        data = self.current_report['data']
        
        # Генерируем HTML-контент в зависимости от типа отчета
        if report_type == 'sites':
            html_content = self._generate_sites_html(data)
        elif report_type == 'changes':
            html_content = self._generate_changes_html(data)
        elif report_type == 'errors':
            html_content = self._generate_errors_html(data)
        elif report_type == 'stats':
            html_content = self._generate_stats_html(data)
        else:
            html_content = "<h1>Неизвестный тип отчета</h1>"
        
        # Устанавливаем HTML-контент в предпросмотр
        self.html_preview.setHtml(html_content)
        
        # Сохраняем сгенерированный контент
        self.current_report['content'] = html_content
    
    def _update_table_preview(self):
        """Обновление табличного предпросмотра"""
        self.logger.debug("Обновление табличного предпросмотра")
        
        # Получаем тип отчета и данные
        report_type = self.current_report['type']
        data = self.current_report['data']
        
        # Очищаем таблицу
        self.table_preview.setRowCount(0)
        
        # Заполняем таблицу в зависимости от типа отчета
        if report_type == 'sites':
            self._fill_sites_table(data)
        elif report_type == 'changes':
            self._fill_changes_table(data)
        elif report_type == 'errors':
            self._fill_errors_table(data)
        elif report_type == 'stats':
            self._fill_stats_table(data)
    
    def _fill_sites_table(self, data):
        """
        Заполнение таблицы данными о сайтах
        
        Args:
            data: Данные о сайтах
        """
        self.logger.debug("Заполнение таблицы данными о сайтах")
        
        # Устанавливаем заголовки столбцов
        self.table_preview.setColumnCount(7)
        self.table_preview.setHorizontalHeaderLabels([
            "ID", "Название", "URL", "Группа", "Последняя проверка", 
            "Последнее изменение", "Статус"
        ])
        
        # Настраиваем растяжение столбцов
        self.table_preview.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeMode.ResizeToContents)
        self.table_preview.horizontalHeader().setSectionResizeMode(1, QHeaderView.ResizeMode.Stretch)
        self.table_preview.horizontalHeader().setSectionResizeMode(2, QHeaderView.ResizeMode.Stretch)
        
        # Заполняем таблицу данными
        for site in data:
            row_position = self.table_preview.rowCount()
            self.table_preview.insertRow(row_position)
            
            # ID сайта
            self.table_preview.setItem(row_position, 0, QTableWidgetItem(str(site.get('id', ''))))
            
            # Название сайта
            self.table_preview.setItem(row_position, 1, QTableWidgetItem(site.get('name', '')))
            
            # URL сайта
            self.table_preview.setItem(row_position, 2, QTableWidgetItem(site.get('url', '')))
            
            # Группа
            self.table_preview.setItem(row_position, 3, QTableWidgetItem(site.get('group', '')))
            
            # Последняя проверка
            last_check = site.get('last_check', '')
            self.table_preview.setItem(row_position, 4, QTableWidgetItem(str(last_check)))
            
            # Последнее изменение
            last_change = site.get('last_change', '')
            self.table_preview.setItem(row_position, 5, QTableWidgetItem(str(last_change)))
            
            # Статус
            status = site.get('status', '')
            status_item = QTableWidgetItem(status)
            self.table_preview.setItem(row_position, 6, status_item)
    
    def _fill_changes_table(self, data):
        """
        Заполнение таблицы данными об изменениях
        
        Args:
            data: Данные об изменениях
        """
        self.logger.debug("Заполнение таблицы данными об изменениях")
        
        # Устанавливаем заголовки столбцов
        self.table_preview.setColumnCount(6)
        self.table_preview.setHorizontalHeaderLabels([
            "ID", "Сайт", "URL", "Дата", "Изменения (%)", "Статус"
        ])
        
        # Настраиваем растяжение столбцов
        self.table_preview.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeMode.ResizeToContents)
        self.table_preview.horizontalHeader().setSectionResizeMode(1, QHeaderView.ResizeMode.Stretch)
        self.table_preview.horizontalHeader().setSectionResizeMode(2, QHeaderView.ResizeMode.Stretch)
        
        # Заполняем таблицу данными
        for change in data:
            row_position = self.table_preview.rowCount()
            self.table_preview.insertRow(row_position)
            
            # ID изменения
            self.table_preview.setItem(row_position, 0, QTableWidgetItem(str(change.get('id', ''))))
            
            # Название сайта
            self.table_preview.setItem(row_position, 1, QTableWidgetItem(change.get('site_name', '')))
            
            # URL сайта
            self.table_preview.setItem(row_position, 2, QTableWidgetItem(change.get('url', '')))
            
            # Дата
            change_date = change.get('date', '')
            self.table_preview.setItem(row_position, 3, QTableWidgetItem(str(change_date)))
            
            # Изменения (%)
            changes_percent = change.get('diff_percent', 0)
            self.table_preview.setItem(row_position, 4, QTableWidgetItem(f"{changes_percent:.2f}%"))
            
            # Статус
            status = change.get('status', '')
            self.table_preview.setItem(row_position, 5, QTableWidgetItem(status))
    
    def _fill_errors_table(self, data):
        """
        Заполнение таблицы данными об ошибках
        
        Args:
            data: Данные об ошибках
        """
        self.logger.debug("Заполнение таблицы данными об ошибках")
        
        # Устанавливаем заголовки столбцов
        self.table_preview.setColumnCount(5)
        self.table_preview.setHorizontalHeaderLabels([
            "ID", "Сайт", "URL", "Дата", "Ошибка"
        ])
        
        # Настраиваем растяжение столбцов
        self.table_preview.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeMode.ResizeToContents)
        self.table_preview.horizontalHeader().setSectionResizeMode(1, QHeaderView.ResizeMode.Stretch)
        self.table_preview.horizontalHeader().setSectionResizeMode(2, QHeaderView.ResizeMode.Stretch)
        self.table_preview.horizontalHeader().setSectionResizeMode(4, QHeaderView.ResizeMode.Stretch)
        
        # Заполняем таблицу данными
        for error in data:
            row_position = self.table_preview.rowCount()
            self.table_preview.insertRow(row_position)
            
            # ID ошибки
            self.table_preview.setItem(row_position, 0, QTableWidgetItem(str(error.get('id', ''))))
            
            # Название сайта
            self.table_preview.setItem(row_position, 1, QTableWidgetItem(error.get('site_name', '')))
            
            # URL сайта
            self.table_preview.setItem(row_position, 2, QTableWidgetItem(error.get('url', '')))
            
            # Дата
            error_date = error.get('date', '')
            self.table_preview.setItem(row_position, 3, QTableWidgetItem(str(error_date)))
            
            # Ошибка
            error_message = error.get('message', '')
            self.table_preview.setItem(row_position, 4, QTableWidgetItem(error_message))
    
    def _fill_stats_table(self, data):
        """
        Заполнение таблицы статистическими данными
        
        Args:
            data: Статистические данные
        """
        self.logger.debug("Заполнение таблицы статистическими данными")
        
        # Устанавливаем заголовки столбцов
        self.table_preview.setColumnCount(2)
        self.table_preview.setHorizontalHeaderLabels([
            "Показатель", "Значение"
        ])
        
        # Настраиваем растяжение столбцов
        self.table_preview.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeMode.Stretch)
        
        # Заполняем таблицу данными
        for key, value in data.items():
            row_position = self.table_preview.rowCount()
            self.table_preview.insertRow(row_position)
            
            # Показатель
            self.table_preview.setItem(row_position, 0, QTableWidgetItem(key))
            
            # Значение
            self.table_preview.setItem(row_position, 1, QTableWidgetItem(str(value)))
    
    def _generate_sites_html(self, data):
        """
        Генерация HTML-таблицы с данными о сайтах
        
        Args:
            data: Данные о сайтах
            
        Returns:
            HTML-код таблицы
        """
        # Получаем период отчета
        date_from = self.date_from.date().toString("dd.MM.yyyy")
        date_to = self.date_to.date().toString("dd.MM.yyyy")
        
        html = f"""
        <!DOCTYPE html>
        <html>
        <head>
            <meta charset="utf-8">
            <title>Отчет по сайтам</title>
            <style>
                body {{ font-family: Arial, sans-serif; margin: 20px; }}
                h1 {{ color: #333366; }}
                table {{ border-collapse: collapse; width: 100%; margin-top: 20px; }}
                th, td {{ text-align: left; padding: 8px; border: 1px solid #ddd; }}
                th {{ background-color: #f2f2f2; color: #333; }}
                tr:nth-child(even) {{ background-color: #f9f9f9; }}
                .header {{ margin-bottom: 20px; }}
                .footer {{ margin-top: 20px; color: #666; font-size: 0.8em; }}
            </style>
        </head>
        <body>
            <div class="header">
                <h1>Отчет по сайтам</h1>
                <p>Период: {date_from} - {date_to}</p>
                <p>Дата создания: {datetime.datetime.now().strftime("%d.%m.%Y %H:%M:%S")}</p>
            </div>
            
            <table>
                <tr>
                    <th>ID</th>
                    <th>Название</th>
                    <th>URL</th>
                    <th>Группа</th>
                    <th>Последняя проверка</th>
                    <th>Последнее изменение</th>
                    <th>Статус</th>
                </tr>
        """
        
        for site in data:
            html += f"""
                <tr>
                    <td>{site.get('id', '')}</td>
                    <td>{site.get('name', '')}</td>
                    <td>{site.get('url', '')}</td>
                    <td>{site.get('group', '')}</td>
                    <td>{site.get('last_check', '')}</td>
                    <td>{site.get('last_change', '')}</td>
                    <td>{site.get('status', '')}</td>
                </tr>
            """
        
        html += """
            </table>
        """
        
        # Добавляем сводную информацию
        total_sites = len(data)
        active_sites = sum(1 for site in data if site.get('status', '') == 'Активен')
        
        html += f"""
            <div class="summary">
                <h3>Сводная информация</h3>
                <p>Всего сайтов: {total_sites}</p>
                <p>Активных сайтов: {active_sites}</p>
                <p>Неактивных сайтов: {total_sites - active_sites}</p>
            </div>
            
            <div class="footer">
                <p>Отчет создан с помощью Web Data Monitor V12</p>
            </div>
        </body>
        </html>
        """
        
        return html
    
    def _generate_changes_html(self, data):
        """
        Генерация HTML-таблицы с данными об изменениях
        
        Args:
            data: Данные об изменениях
            
        Returns:
            HTML-код таблицы
        """
        # Получаем период отчета
        date_from = self.date_from.date().toString("dd.MM.yyyy")
        date_to = self.date_to.date().toString("dd.MM.yyyy")
        
        html = f"""
        <!DOCTYPE html>
        <html>
        <head>
            <meta charset="utf-8">
            <title>Отчет по изменениям</title>
            <style>
                body {{ font-family: Arial, sans-serif; margin: 20px; }}
                h1 {{ color: #333366; }}
                table {{ border-collapse: collapse; width: 100%; margin-top: 20px; }}
                th, td {{ text-align: left; padding: 8px; border: 1px solid #ddd; }}
                th {{ background-color: #f2f2f2; color: #333; }}
                tr:nth-child(even) {{ background-color: #f9f9f9; }}
                .header {{ margin-bottom: 20px; }}
                .footer {{ margin-top: 20px; color: #666; font-size: 0.8em; }}
            </style>
        </head>
        <body>
            <div class="header">
                <h1>Отчет по изменениям</h1>
                <p>Период: {date_from} - {date_to}</p>
                <p>Дата создания: {datetime.datetime.now().strftime("%d.%m.%Y %H:%M:%S")}</p>
            </div>
            
            <table>
                <tr>
                    <th>ID</th>
                    <th>Сайт</th>
                    <th>URL</th>
                    <th>Дата</th>
                    <th>Изменения (%)</th>
                    <th>Статус</th>
                </tr>
        """
        
        for change in data:
            diff_percent = change.get('diff_percent', 0)
            html += f"""
                <tr>
                    <td>{change.get('id', '')}</td>
                    <td>{change.get('site_name', '')}</td>
                    <td>{change.get('url', '')}</td>
                    <td>{change.get('date', '')}</td>
                    <td>{diff_percent:.2f}%</td>
                    <td>{change.get('status', '')}</td>
                </tr>
            """
        
        html += """
            </table>
        """
        
        # Добавляем сводную информацию
        total_changes = len(data)
        avg_diff = sum(change.get('diff_percent', 0) for change in data) / max(total_changes, 1)
        
        html += f"""
            <div class="summary">
                <h3>Сводная информация</h3>
                <p>Всего изменений: {total_changes}</p>
                <p>Среднее изменение: {avg_diff:.2f}%</p>
            </div>
            
            <div class="footer">
                <p>Отчет создан с помощью Web Data Monitor V12</p>
            </div>
        </body>
        </html>
        """
        
        return html
    
    def _generate_errors_html(self, data):
        """
        Генерация HTML-таблицы с данными об ошибках
        
        Args:
            data: Данные об ошибках
            
        Returns:
            HTML-код таблицы
        """
        # Получаем период отчета
        date_from = self.date_from.date().toString("dd.MM.yyyy")
        date_to = self.date_to.date().toString("dd.MM.yyyy")
        
        html = f"""
        <!DOCTYPE html>
        <html>
        <head>
            <meta charset="utf-8">
            <title>Отчет по ошибкам</title>
            <style>
                body {{ font-family: Arial, sans-serif; margin: 20px; }}
                h1 {{ color: #333366; }}
                table {{ border-collapse: collapse; width: 100%; margin-top: 20px; }}
                th, td {{ text-align: left; padding: 8px; border: 1px solid #ddd; }}
                th {{ background-color: #f2f2f2; color: #333; }}
                tr:nth-child(even) {{ background-color: #f9f9f9; }}
                .header {{ margin-bottom: 20px; }}
                .footer {{ margin-top: 20px; color: #666; font-size: 0.8em; }}
            </style>
        </head>
        <body>
            <div class="header">
                <h1>Отчет по ошибкам</h1>
                <p>Период: {date_from} - {date_to}</p>
                <p>Дата создания: {datetime.datetime.now().strftime("%d.%m.%Y %H:%M:%S")}</p>
            </div>
            
            <table>
                <tr>
                    <th>ID</th>
                    <th>Сайт</th>
                    <th>URL</th>
                    <th>Дата</th>
                    <th>Ошибка</th>
                </tr>
        """
        
        for error in data:
            html += f"""
                <tr>
                    <td>{error.get('id', '')}</td>
                    <td>{error.get('site_name', '')}</td>
                    <td>{error.get('url', '')}</td>
                    <td>{error.get('date', '')}</td>
                    <td>{error.get('message', '')}</td>
                </tr>
            """
        
        html += """
            </table>
        """
        
        # Добавляем сводную информацию
        total_errors = len(data)
        
        html += f"""
            <div class="summary">
                <h3>Сводная информация</h3>
                <p>Всего ошибок: {total_errors}</p>
            </div>
            
            <div class="footer">
                <p>Отчет создан с помощью Web Data Monitor V12</p>
            </div>
        </body>
        </html>
        """
        
        return html
    
    def _generate_stats_html(self, data):
        """
        Генерация HTML-таблицы со статистическими данными
        
        Args:
            data: Статистические данные
            
        Returns:
            HTML-код таблицы
        """
        # Получаем период отчета
        date_from = self.date_from.date().toString("dd.MM.yyyy")
        date_to = self.date_to.date().toString("dd.MM.yyyy")
        
        html = f"""
        <!DOCTYPE html>
        <html>
        <head>
            <meta charset="utf-8">
            <title>Статистический отчет</title>
            <style>
                body {{ font-family: Arial, sans-serif; margin: 20px; }}
                h1 {{ color: #333366; }}
                table {{ border-collapse: collapse; width: 100%; margin-top: 20px; }}
                th, td {{ text-align: left; padding: 8px; border: 1px solid #ddd; }}
                th {{ background-color: #f2f2f2; color: #333; }}
                tr:nth-child(even) {{ background-color: #f9f9f9; }}
                .header {{ margin-bottom: 20px; }}
                .footer {{ margin-top: 20px; color: #666; font-size: 0.8em; }}
            </style>
        </head>
        <body>
            <div class="header">
                <h1>Статистический отчет</h1>
                <p>Период: {date_from} - {date_to}</p>
                <p>Дата создания: {datetime.datetime.now().strftime("%d.%m.%Y %H:%M:%S")}</p>
            </div>
            
            <table>
                <tr>
                    <th>Показатель</th>
                    <th>Значение</th>
                </tr>
        """
        
        for key, value in data.items():
            html += f"""
                <tr>
                    <td>{key}</td>
                    <td>{value}</td>
                </tr>
            """
        
        html += """
            </table>
            
            <div class="footer">
                <p>Отчет создан с помощью Web Data Monitor V12</p>
            </div>
        </body>
        </html>
        """
        
        return html
    
    def _on_print_report(self):
        """Обработчик печати отчета"""
        self.logger.debug("Вызван метод печати отчета")
        
        try:
            # Создаем принтер
            printer = QPrinter(QPrinter.PrinterMode.HighResolution)
            
            # Открываем диалог печати
            dialog = QPrintDialog(printer, self)
            if dialog.exec() != QPrintDialog.DialogCode.Accepted:
                return
            
            # Создаем документ
            document = QTextDocument()
            document.setHtml(self.current_report['content'])
            
            # Печатаем документ
            document.print(printer)
            
            self.logger.info("Отчет успешно отправлен на печать")
            QMessageBox.information(self, "Успех", "Отчет успешно отправлен на печать")
        
        except Exception as e:
            self.logger.error(f"Ошибка при печати отчета: {e}")
            log_exception(self.logger, "Ошибка печати отчета")
            QMessageBox.critical(self, "Ошибка", f"Не удалось напечатать отчет: {e}") 